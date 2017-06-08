#!/usr/bin/env python


import argparse
from openpyxl import load_workbook


class GetItems:
    def __init__(self, file_name):
       wb = load_workbook(filename = file_name)
       self.sheet = wb.get_active_sheet()

    def parse_sheet(self):
        for row in range(2, self.sheet.max_row + 1):
            date = self.sheet['A' + str(row)].value
            fruit = self.sheet['B' + str(row)].value
            sample_number = self.sheet['C' + str(row)].value
            print date, fruit, sample_number
            print


if __name__=="__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file', dest='file_path', type=file)
    results = parser.parse_args()

    obj = GetItems(results.file_path)

    obj.parse_sheet()


